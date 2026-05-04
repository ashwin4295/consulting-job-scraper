#!/usr/bin/env python3
"""
Weekly Consulting Job Board Scraper
Scrapes jobs from MyConsultingOffer and Management Consulted job boards,
then generates a McKinsey-styled Excel file.

Requirements: playwright, openpyxl
Install: pip install playwright openpyxl && playwright install chromium
"""

import time, os, concurrent.futures
from datetime import date
from urllib.parse import urlparse, parse_qs, unquote
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


MCO_URL = "https://jobs.myconsultingoffer.org/jobs"
MC_URL = "https://jobs.managementconsulted.com/jobs"
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", ".")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"consulting_job_board_{date.today().isoformat()}.xlsx")

COMPANY_WEBSITES = {
    'McKinsey & Company': 'https://www.mckinsey.com',
    'Boston Consulting Group (BCG)': 'https://www.bcg.com',
    'Boston Consulting Group': 'https://www.bcg.com',
    'Bain & Company': 'https://www.bain.com',
    'Roland Berger': 'https://www.rolandberger.com',
    'Oliver Wyman': 'https://www.oliverwyman.com',
    'Alvarez & Marsal': 'https://www.alvarezandmarsal.com',
    'EY': 'https://www.ey.com',
    'EY-Parthenon': 'https://www.ey.com/en_gl/services/strategy/ey-parthenon',
    'Deloitte': 'https://www.deloitte.com',
    'KPMG UK': 'https://www.kpmg.co.uk',
    'KPMG Canada': 'https://home.kpmg/ca',
    'KPMG US': 'https://www.kpmg.us',
    'PwC': 'https://www.pwc.com',
    'PwC Canada': 'https://www.pwc.com/ca',
    'Accenture': 'https://www.accenture.com',
    'Capital One': 'https://www.capitalone.com',
    'Google': 'https://careers.google.com',
    'Guidehouse': 'https://guidehouse.com',
    'Huron': 'https://www.huronconsultinggroup.com',
    'Simon-Kucher': 'https://www.simon-kucher.com',
    'FTI Consulting': 'https://www.fticonsulting.com',
    'Booz Allen Hamilton': 'https://www.boozallen.com',
    'L.E.K. Consulting': 'https://www.lek.com',
    'Kearney': 'https://www.kearney.com',
    'AlixPartners': 'https://www.alixpartners.com',
    'Infosys': 'https://www.infosys.com',
    'Cognizant': 'https://www.cognizant.com',
    'IQVIA': 'https://www.iqvia.com',
    'Visa': 'https://www.visa.com',
    'The Walt Disney Company': 'https://thewaltdisneycompany.com',
    'American Express': 'https://www.americanexpress.com',
    'Ford Motor Company': 'https://www.ford.com',
    'General Motors': 'https://www.gm.com',
    'Pfizer': 'https://www.pfizer.com',
    'The Carlyle Group': 'https://www.carlyle.com',
    'RBC': 'https://www.rbc.com',
    'Trinity Life Sciences': 'https://trinitylifesciences.com',
    'Mercer': 'https://www.mercer.com',
    'West Monroe': 'https://www.westmonroe.com',
    'Stax': 'https://www.stax.com',
    'Porsche Consulting': 'https://www.porsche-consulting.com',
    'RSM Canada': 'https://www.rsmcanada.com',
    'The Coca-Cola Company': 'https://www.coca-colacompany.com',
    'Kantar': 'https://www.kantar.com',
    'ClearView Healthcare Partners': 'https://www.clearviewhcp.com',
    'Clarkston Consulting': 'https://clarkstonconsulting.com',
    'North Highland': 'https://www.northhighland.com',
}


# Firm-specific careers landing pages, used as a fallback when a job has no
# external apply URL on the source board. Only firms with a known, stable
# careers URL are listed here; anything else gets a blank apply cell.
FIRM_CAREERS = {
    'McKinsey & Company': 'https://www.mckinsey.com/careers/search-jobs',
    'Boston Consulting Group (BCG)': 'https://careers.bcg.com',
    'Boston Consulting Group': 'https://careers.bcg.com',
    'Bain & Company': 'https://www.bain.com/careers/find-a-role/',
    'Roland Berger': 'https://www.rolandberger.com/en/Careers/',
    'Oliver Wyman': 'https://www.oliverwyman.com/careers.html',
    'Alvarez & Marsal': 'https://careers.alvarezandmarsal.com',
    'Alvarez and Marsal': 'https://careers.alvarezandmarsal.com',
    'EY': 'https://careers.ey.com',
    'EY-Parthenon': 'https://careers.ey.com',
    'Deloitte': 'https://apply.deloitte.com',
    'KPMG UK': 'https://www.kpmgcareers.co.uk',
    'KPMG Canada': 'https://kpmg.com/ca/en/home/careers.html',
    'KPMG US': 'https://www.kpmgus.com/careers.html',
    'PwC': 'https://jobs.us.pwc.com',
    'PwC Canada': 'https://www.pwc.com/ca/en/careers.html',
    'Accenture': 'https://www.accenture.com/us-en/careers',
    'Capital One': 'https://www.capitalonecareers.com',
    'Google': 'https://careers.google.com/jobs/results/',
    'Guidehouse': 'https://guidehouse.wd1.myworkdayjobs.com/External',
    'Huron': 'https://huron.wd1.myworkdayjobs.com/HuronCareers',
    'Simon-Kucher': 'https://www.simon-kucher.com/en/careers',
    'FTI Consulting': 'https://www.fticonsulting.com/careers',
    'Booz Allen Hamilton': 'https://www.boozallen.com/careers.html',
    'L.E.K. Consulting': 'https://www.lek.com/careers',
    'Kearney': 'https://www.kearney.com/careers',
    'AlixPartners': 'https://www.alixpartners.com/careers/',
    'Infosys': 'https://www.infosys.com/careers/',
    'Cognizant': 'https://careers.cognizant.com',
    'IQVIA': 'https://jobs.iqvia.com',
    'Visa': 'https://corporate.visa.com/en/jobs.html',
    'The Walt Disney Company': 'https://jobs.disneycareers.com',
    'American Express': 'https://www.americanexpress.com/en-us/careers',
    'Ford Motor Company': 'https://corporate.ford.com/careers.html',
    'General Motors': 'https://search-careers.gm.com',
    'Pfizer': 'https://www.pfizer.com/about/careers',
    'The Carlyle Group': 'https://www.carlyle.com/careers',
    'RBC': 'https://jobs.rbc.com',
    'Trinity Life Sciences': 'https://trinitylifesciences.com/careers/',
    'Mercer': 'https://careers.marshmclennan.com',
    'West Monroe': 'https://www.westmonroe.com/careers',
    'Stax': 'https://www.stax.com/careers',
    'Porsche Consulting': 'https://www.porsche-consulting.com/en/career',
    'RSM Canada': 'https://rsmcanada.com/careers.html',
    'The Coca-Cola Company': 'https://careers.coca-colacompany.com',
    'Kantar': 'https://www.kantar.com/careers',
    'ClearView Healthcare Partners': 'https://www.clearviewhcp.com/careers',
    'Clarkston Consulting': 'https://clarkstonconsulting.com/careers/',
    'North Highland': 'https://www.northhighland.com/careers',
    'Teneo': 'https://www.teneo.com/careers/',
    'Synpulse': 'https://www.synpulse.com/en/careers',
    'PA Consulting': 'https://www.paconsulting.com/careers',
    'TPG': 'https://www.tpg.com/careers',
}


_DDG_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def _ddg_search(query, max_results=5, timeout=10):
    """Hit DuckDuckGo's HTML interface and return result URLs in rank order."""
    try:
        resp = requests.post(
            "https://html.duckduckgo.com/html/",
            data={"q": query},
            headers=_DDG_HEADERS,
            timeout=timeout,
        )
        resp.raise_for_status()
    except Exception:
        return []
    soup = BeautifulSoup(resp.text, "html.parser")
    out = []
    for a in soup.select("a.result__a"):
        href = a.get("href", "")
        if not href:
            continue
        # DDG wraps results as /l/?uddg=<encoded-url>
        if href.startswith("//duckduckgo.com/l/") or href.startswith("/l/"):
            wrapped = "https:" + href if href.startswith("//") else "https://duckduckgo.com" + href
            uddg = parse_qs(urlparse(wrapped).query).get("uddg", [""])[0]
            if uddg:
                href = unquote(uddg)
        if href.startswith("http"):
            out.append(href)
        if len(out) >= max_results:
            break
    return out


def find_firm_apply_url(company, title):
    """Return a firm-specific apply URL for a job that has no external apply URL.

    Strategy:
      1. Look up the firm's careers landing page in FIRM_CAREERS. If we don't
         have a stable URL for this firm, give up — better blank than wrong.
      2. Search DuckDuckGo constrained to that careers domain for the specific
         job title; if a match comes back, return it.
      3. Otherwise return the firm's general careers URL.
    """
    careers_url = FIRM_CAREERS.get(company or "")
    if not careers_url:
        return ""
    domain = urlparse(careers_url).netloc
    if domain:
        try:
            results = _ddg_search(
                f'"{title}" "{company}" site:{domain}', max_results=5, timeout=10
            )
            for r in results:
                if domain in urlparse(r).netloc:
                    return r
        except Exception:
            pass
    return careers_url


def enrich_blank_apply_links(jobs, max_searches=200, max_workers=8):
    """For jobs with no applyLink, fill in a firm-specific careers link.

    Specific job URLs (via search) are tried first, capped at `max_searches`
    to keep the weekly cron fast and avoid DDG rate limiting. Beyond that cap,
    we fall straight to the firm's general careers URL with no search.
    """
    blanks = [j for j in jobs if not j.get("applyLink")]
    if not blanks:
        return 0

    print(f"\nEnriching {len(blanks)} jobs without external apply URL...")

    # Cache by (company, title) so duplicate roles don't search twice.
    cache = {}
    searchable = blanks[:max_searches]
    direct = blanks[max_searches:]

    def resolve(job):
        key = (job.get("company", ""), job.get("title", ""))
        if key in cache:
            return job, cache[key]
        url = find_firm_apply_url(job.get("company", ""), job.get("title", ""))
        cache[key] = url
        return job, url

    filled = 0
    if searchable:
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
            for job, url in ex.map(resolve, searchable):
                if url:
                    job["applyLink"] = url
                    filled += 1
    for job in direct:
        # Skip the search, just use the firm careers page if known.
        url = FIRM_CAREERS.get(job.get("company", ""), "")
        if url:
            job["applyLink"] = url
            filled += 1
    print(f"  filled {filled}/{len(blanks)} blanks")
    return filled


CARD_EXTRACTOR_JS = r"""
() => {
    // Each Jboard listing page exposes the full job payload as window.jobsList.
    // This is the source of truth — including apply_by/apply_to, which give us
    // the real employer ATS URL when the employer chose external-link applies.
    const list = Array.isArray(window.jobsList) ? window.jobsList : [];
    const tenantHost = location.host;
    const out = [];
    for (const j of list) {
        if (!j || !j.id || !j.title) continue;

        // Only treat apply_to as a real external link when the employer opted into
        // by-link applies AND the URL leaves the Jboard tenant. Anything else
        // (in-board form, email, missing URL, or Jboard-hosted URL) is dropped.
        let applyLink = '';
        if (j.apply_by === 'by_link' && typeof j.apply_to === 'string' && j.apply_to) {
            try {
                const u = new URL(j.apply_to);
                if (u.host && u.host !== tenantHost && !/jboard\.io$/.test(u.host)) {
                    applyLink = j.apply_to;
                }
            } catch (e) { /* malformed URL, skip */ }
        }

        out.push({
            jobId: String(j.id),
            title: j.title || '',
            company: (j.employer && j.employer.name) || '',
            location: j.location || '',
            jobType: (j.job_type && j.job_type.name) || '',
            category: (j.category && j.category.name) || '',
            slug: j.job_details_path || '',
            applyLink,
        });
    }
    return out;
}
"""


def scrape_jboard_site(page, base_url, site_name):
    print(f"\nScraping {site_name}: {base_url}")
    all_jobs = []
    seen_ids = set()
    page_num = 1
    max_pages = 100  # safety cap

    while page_num <= max_pages:
        url = base_url if page_num == 1 else f"{base_url}?page={page_num}"
        print(f"  Page {page_num}: {url}")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except Exception as e:
            print(f"  goto failed: {e}; retrying with networkidle...")
            page.goto(url, wait_until="networkidle", timeout=60000)

        loaded = False
        for attempt in range(3):
            try:
                page.wait_for_selector(".job-listings-item", timeout=20000)
                loaded = True
                break
            except Exception:
                title = page.title()
                if "moment" in title.lower() or "challenge" in title.lower():
                    print(f"  Cloudflare challenge detected (title={title!r}), waiting and retrying...")
                    time.sleep(8)
                    continue
                break
        if not loaded:
            print("  No .job-listings-item on page after retries -- stopping.")
            break

        time.sleep(1)
        jobs = page.evaluate(CARD_EXTRACTOR_JS)
        new_count = 0
        for j in jobs:
            jid = j.get("jobId", "")
            if jid and jid not in seen_ids:
                seen_ids.add(jid)
                all_jobs.append(j)
                new_count += 1
        print(f"    extracted {len(jobs)} cards ({new_count} new), total so far: {len(all_jobs)}")

        next_href = page.evaluate(
            "(()=>{const l=document.querySelector('link[rel=next]');return l?l.getAttribute('href'):null})()"
        )
        if not next_href:
            print("  No rel=next link — done.")
            break
        if new_count == 0:
            print("  No new jobs on this page — stopping to avoid loop.")
            break
        page_num += 1

    print(f"Extracted {len(all_jobs)} unique jobs from {site_name}")
    return all_jobs


def scrape_site(page, site_url, site_name):
    return scrape_jboard_site(page, site_url, site_name)


def build_excel(all_jobs, output_path):
    print(f"\nBuilding Excel with {len(all_jobs)} jobs...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Board"

    MCK_NAVY, MCK_BLUE = '003A5C', '0072CE'
    MCK_LIGHT_BLUE, MCK_LIGHT_GRAY = 'E8F4FD', 'F5F5F5'
    MCK_WHITE, MCK_BORDER_COLOR = 'FFFFFF', 'D0D0D0'

    header_fill = PatternFill('solid', fgColor=MCK_NAVY)
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=9.5, color='333333')
    link_font = Font(name='Arial', size=9.5, color=MCK_BLUE, underline='single')
    data_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color=MCK_BORDER_COLOR),
        right=Side(style='thin', color=MCK_BORDER_COLOR),
        top=Side(style='thin', color=MCK_BORDER_COLOR),
        bottom=Side(style='thin', color=MCK_BORDER_COLOR),
    )
    alt_fill = PatternFill('solid', fgColor=MCK_LIGHT_GRAY)
    white_fill = PatternFill('solid', fgColor=MCK_WHITE)

    ws.merge_cells('A1:H1')
    ws['A1'].value = 'MBB Prep - Management Consulting Job Board'
    ws['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = PatternFill('solid', fgColor=MCK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:H2')
    ws['A2'].value = f'{len(all_jobs)} consulting jobs  |  Updated {date.today().strftime("%b %d, %Y")}'
    ws['A2'].font = Font(name='Arial', size=10, color='666666', italic=True)
    ws['A2'].fill = PatternFill('solid', fgColor=MCK_LIGHT_BLUE)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 8

    headers = ['#', 'Job Title', 'Firm', 'Category', 'Job Type', 'Location', 'Firm Website', 'Application Link']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    for i, job in enumerate(all_jobs):
        row = i + 5
        row_fill = alt_fill if i % 2 == 0 else white_fill
        ws.cell(row=row, column=1, value=i+1).font = Font(name='Arial', size=9, color='999999')
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=2, value=job['title']).font = Font(name='Arial', size=9.5, color='333333', bold=True)
        ws.cell(row=row, column=2).alignment = data_align
        ws.cell(row=row, column=3, value=job['company']).font = data_font
        ws.cell(row=row, column=3).alignment = data_align
        cat = job.get('category', '') or 'Management Consulting'
        ws.cell(row=row, column=4, value=cat).font = data_font
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=5, value=job.get('jobType', '')).font = data_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=6, value=job.get('location', '')).font = data_font
        ws.cell(row=row, column=6).alignment = data_align
        website = job.get('website', '')
        if website:
            domain = website.replace('https://', '').replace('http://', '').split('/')[0]
            c = ws.cell(row=row, column=7, value=domain)
            c.font = link_font
            c.hyperlink = website
        ws.cell(row=row, column=7).alignment = data_align
        apply_url = job.get('applyLink', '')
        if apply_url:
            c = ws.cell(row=row, column=8, value='Apply')
            c.font = link_font
            c.hyperlink = apply_url
        ws.cell(row=row, column=8).alignment = data_align
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = row_fill
            ws.cell(row=row, column=col).border = thin_border

    widths = {'A': 5, 'B': 55, 'C': 28, 'D': 22, 'E': 12, 'F': 35, 'G': 30, 'H': 12}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.auto_filter.ref = f'A4:H{4 + len(all_jobs)}'
    ws.freeze_panes = 'A5'
    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    all_jobs = []
    seen_ids = set()

    with Stealth().use_sync(sync_playwright()) as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )
        context = browser.new_context(
            viewport={"width": 1366, "height": 900},
            locale="en-US",
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            },
        )

        for board_url, board_name in [(MCO_URL, "MyConsultingOffer"), (MC_URL, "Management Consulted")]:
            board_page = context.new_page()
            jobs = scrape_site(board_page, board_url, board_name)
            for j in jobs:
                jid = j.get('jobId', '')
                if jid and jid not in seen_ids:
                    seen_ids.add(jid)
                    j['website'] = COMPANY_WEBSITES.get(j['company'], '')
                    # j['applyLink'] is already set by the extractor — it is either
                    # the real employer ATS URL (apply_by=by_link, off-tenant) or ''.
                    # Never substitute the board's own /apply URL: that's a competitor.
                    all_jobs.append(j)
            board_page.close()
        browser.close()

    # For any remaining blank applyLinks, fall back to a firm-specific URL:
    # specific job posting on the firm's careers site if findable, else the
    # firm's general careers landing page. Still leaves blanks empty when the
    # firm isn't in FIRM_CAREERS.
    enrich_blank_apply_links(all_jobs)

    print(f"\nTotal unique jobs: {len(all_jobs)}")
    if all_jobs:
        build_excel(all_jobs, OUTPUT_FILE)
    else:
        print("ERROR: No jobs scraped. Check if site structure changed.")
        exit(1)


if __name__ == "__main__":
    main()
